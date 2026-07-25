from __future__ import annotations

import argparse
import json
import math
import multiprocessing as mp
import os
import signal
import sys
import time
from dataclasses import asdict
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook

os.environ.setdefault("OMP_NUM_THREADS", "1")
os.environ.setdefault("OPENBLAS_NUM_THREADS", "1")
os.environ.setdefault("MKL_NUM_THREADS", "1")

PROJECT_ROOT = Path(__file__).resolve().parent
CAMOUFLAGE_ROOT = Path("/home/jushuo/Code/zz3-3D-camouflage")
DEFAULT_OUTPUT = PROJECT_ROOT / "output" / "target_prediction_attack_sweep.xlsx"
DEFAULT_SAMPLES = [
    "sample-292",
    "sample-133",
    "sample-045",
    "sample-002",
    "sample-071",
    "sample-061",
    "sample-192",
    "sample-001",
]

sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(CAMOUFLAGE_ROOT))

from frenet_interface import CartesianEgoState, FrenetPlannerConfig, ObstaclePoint, plan_once  # noqa: E402
from nuscenes_fot import NuScenesFOTAdapter, NuScenesFOTAdapterConfig, NuScenesFOTPaths  # noqa: E402
from match_target_car import run_target_matching, _save_yaml as _save_match_yaml  # noqa: E402


_WORKER_PLANNER_CONFIG: FrenetPlannerConfig | None = None
_WORKER_PLAN_TIMEOUT_S: float = 10.0


def _worker_init(planner_config_values: dict[str, Any], plan_timeout_s: float = 10.0) -> None:
    global _WORKER_PLANNER_CONFIG
    global _WORKER_PLAN_TIMEOUT_S
    _WORKER_PLANNER_CONFIG = FrenetPlannerConfig(**planner_config_values)
    _WORKER_PLAN_TIMEOUT_S = float(plan_timeout_s)


class _PlanTimeout(Exception):
    pass


def _handle_plan_timeout(_signum: int, _frame: Any) -> None:
    raise _PlanTimeout("plan_once timeout")


def _plan_once_with_timeout(*args: Any, timeout_s: float, **kwargs: Any) -> Any:
    if timeout_s <= 0.0:
        return plan_once(*args, **kwargs)
    previous_handler = signal.signal(signal.SIGALRM, _handle_plan_timeout)
    signal.setitimer(signal.ITIMER_REAL, float(timeout_s))
    try:
        return plan_once(*args, **kwargs)
    finally:
        signal.setitimer(signal.ITIMER_REAL, 0.0)
        signal.signal(signal.SIGALRM, previous_handler)


def _load_config(path: Path) -> dict[str, Any]:
    with path.open() as handle:
        return yaml.safe_load(handle) or {}


def _latest_binding_path(
    sample_name: str,
    cache_root: Path,
    *,
    sequence_yaml: Path | None = None,
    match_config_path: Path | None = None,
) -> Path:
    candidates = sorted(
        (cache_root / sample_name).glob("cache-*/target-car-binding.yaml"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if candidates:
        return candidates[0]
    if sequence_yaml is None or match_config_path is None:
        raise FileNotFoundError(f"no target-car-binding.yaml found for {sample_name}")

    binding, _bound_sequence, _sequence_yaml_path = run_target_matching(
        config_path=match_config_path,
        sequence_yaml=sequence_yaml,
        near_plane_m=0.1,
        verbose=False,
    )
    output_path = cache_root / sample_name / "cache-planning-binding" / "target-car-binding.yaml"
    _save_match_yaml(output_path, binding)
    return output_path


def _find_sequence_yaml(sample_name: str, yaml_root: Path) -> Path:
    matches = sorted(yaml_root.glob(f"**/{sample_name}.yaml"))
    if not matches:
        raise FileNotFoundError(f"sequence yaml not found for {sample_name} under {yaml_root}")
    if len(matches) > 1:
        # Prefer exact case-specific selection when callers pass explicit samples from scenario roots.
        return matches[0]
    return matches[0]


def _binding_worker(task: dict[str, str]) -> dict[str, str]:
    sample_name = task["sample_name"]
    sequence_yaml = Path(task["sequence_yaml"])
    cache_root = Path(task["cache_root"])
    match_config_path = Path(task["match_config_path"])
    output_path = cache_root / sample_name / "cache-planning-binding" / "target-car-binding.yaml"
    try:
        if output_path.exists():
            return {"sample_name": sample_name, "path": str(output_path), "error": ""}
        binding, _bound_sequence, _sequence_yaml_path = run_target_matching(
            config_path=match_config_path,
            sequence_yaml=sequence_yaml,
            near_plane_m=0.1,
            verbose=False,
        )
        _save_match_yaml(output_path, binding)
        return {"sample_name": sample_name, "path": str(output_path), "error": ""}
    except Exception as exc:
        return {
            "sample_name": sample_name,
            "path": str(output_path),
            "error": f"{type(exc).__name__}: {exc}",
        }


def _ensure_binding_cache_parallel(
    sample_names: list[str],
    *,
    cache_root: Path,
    yaml_root: Path,
    match_config_path: Path,
    workers: int,
) -> list[dict[str, str]]:
    tasks: list[dict[str, str]] = []
    for sample_name in sample_names:
        existing = sorted((cache_root / sample_name).glob("cache-*/target-car-binding.yaml"))
        if existing:
            continue
        sequence_yaml = _find_sequence_yaml(sample_name, yaml_root)
        tasks.append(
            {
                "sample_name": sample_name,
                "sequence_yaml": str(sequence_yaml),
                "cache_root": str(cache_root),
                "match_config_path": str(match_config_path),
            }
        )
    if not tasks:
        return []
    worker_count = max(1, min(int(workers), len(tasks)))
    print(f"[binding] missing={len(tasks)} workers={worker_count}", flush=True)
    errors: list[dict[str, str]] = []
    if worker_count <= 1:
        for index, task in enumerate(tasks, start=1):
            result = _binding_worker(task)
            if result.get("error"):
                errors.append(result)
            if index % 10 == 0 or index == len(tasks):
                print(f"[binding] {index}/{len(tasks)}", flush=True)
        return errors
    with mp.Pool(processes=worker_count) as pool:
        for index, result in enumerate(pool.imap_unordered(_binding_worker, tasks, chunksize=1), start=1):
            if result.get("error"):
                errors.append(result)
            if index % 10 == 0 or index == len(tasks):
                print(f"[binding] {index}/{len(tasks)} errors={len(errors)}", flush=True)
    return errors


def _load_bindings(
    sample_names: list[str],
    cache_root: Path,
    *,
    yaml_root: Path,
    match_config_path: Path,
) -> dict[str, dict[str, Any]]:
    bindings: dict[str, dict[str, Any]] = {}
    for sample_name in sample_names:
        sequence_yaml = _find_sequence_yaml(sample_name, yaml_root)
        path = _latest_binding_path(
            sample_name,
            cache_root,
            sequence_yaml=sequence_yaml,
            match_config_path=match_config_path,
        )
        with path.open() as handle:
            payload = yaml.safe_load(handle) or {}
        payload["_binding_path"] = str(path)
        payload["_case_name"] = str(sequence_yaml.parent.name)
        bindings[sample_name] = payload
    return bindings


def _discover_scenario_cases(scenario_root: Path) -> dict[str, list[str]]:
    cases: dict[str, list[str]] = {}
    for case_dir in sorted([p for p in scenario_root.iterdir() if p.is_dir()]):
        samples = sorted(
            p.name for p in case_dir.iterdir() if p.is_dir() and p.name.startswith("sample-")
        )
        if samples:
            cases[case_dir.name] = samples
    if not cases:
        raise FileNotFoundError(f"no case/sample directories found under {scenario_root}")
    return cases


def _load_json(path: Path) -> Any:
    with path.open() as handle:
        return json.load(handle)


def _build_gt_tracking_json(
    *,
    nuscenes_root: Path,
    sample_tokens: list[str],
    output_path: Path,
) -> Path:
    version_root = nuscenes_root / "v1.0-trainval"
    needed = set(sample_tokens)
    instances = {row["token"]: row for row in _load_json(version_root / "instance.json")}
    categories = {row["token"]: row["name"] for row in _load_json(version_root / "category.json")}
    results: dict[str, list[dict[str, Any]]] = {token: [] for token in sample_tokens}

    for ann in _load_json(version_root / "sample_annotation.json"):
        sample_token = str(ann.get("sample_token", ""))
        if sample_token not in needed:
            continue
        instance = instances.get(str(ann.get("instance_token", "")), {})
        category_name = categories.get(str(instance.get("category_token", "")), "")
        if not str(category_name).startswith("vehicle"):
            continue
        results[sample_token].append(
            {
                "sample_token": sample_token,
                "translation": ann["translation"],
                "size": ann["size"],
                "rotation": ann["rotation"],
                "tracking_id": ann["instance_token"],
                "tracking_name": category_name.split(".")[-1],
                "tracking_score": 1.0,
            }
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w") as handle:
        json.dump({"results": results, "meta": {"source": "gt_sample_annotation"}}, handle, indent=2)
    return output_path


def _ego_local_xy(global_xy: tuple[float, float], ego: CartesianEgoState) -> tuple[float, float]:
    dx = float(global_xy[0]) - float(ego.x)
    dy = float(global_xy[1]) - float(ego.y)
    cos_yaw = math.cos(float(ego.yaw))
    sin_yaw = math.sin(float(ego.yaw))
    front = cos_yaw * dx + sin_yaw * dy
    left = -sin_yaw * dx + cos_yaw * dy
    return front, left


def _ego_direction_to_global(
    *,
    ego: CartesianEgoState,
    forward_weight: float,
    lateral_weight: float,
    lateral_sign: float,
) -> tuple[float, float]:
    cos_yaw = math.cos(float(ego.yaw))
    sin_yaw = math.sin(float(ego.yaw))
    gx = forward_weight * cos_yaw + lateral_sign * lateral_weight * (-sin_yaw)
    gy = forward_weight * sin_yaw + lateral_sign * lateral_weight * cos_yaw
    norm = math.hypot(gx, gy)
    if norm <= 1.0e-9:
        return -sin_yaw, cos_yaw
    return gx / norm, gy / norm


def _extend_clean_future_points(
    centers_xy: list[tuple[float, float]],
    frame_index: int,
    prediction_steps: int,
) -> list[tuple[float, float]]:
    future = list(centers_xy[frame_index + 1 :])
    if frame_index + 1 < len(centers_xy):
        vx = centers_xy[frame_index + 1][0] - centers_xy[frame_index][0]
        vy = centers_xy[frame_index + 1][1] - centers_xy[frame_index][1]
    elif frame_index > 0:
        vx = centers_xy[frame_index][0] - centers_xy[frame_index - 1][0]
        vy = centers_xy[frame_index][1] - centers_xy[frame_index - 1][1]
    else:
        vx = 0.0
        vy = 0.0

    last = future[-1] if future else centers_xy[frame_index]
    while len(future) < prediction_steps:
        last = (last[0] + vx, last[1] + vy)
        future.append(last)
    return future[:prediction_steps]


def _fake_future_points(
    *,
    current_xy: tuple[float, float],
    direction_xy: tuple[float, float],
    interval_m: float,
    prediction_steps: int,
) -> list[tuple[float, float]]:
    return [
        (
            current_xy[0] + (idx + 1) * interval_m * direction_xy[0],
            current_xy[1] + (idx + 1) * interval_m * direction_xy[1],
        )
        for idx in range(prediction_steps)
    ]


def _prediction_points_to_obstacles(points_xy: list[tuple[float, float]]) -> list[ObstaclePoint]:
    return [ObstaclePoint(x=float(x), y=float(y)) for x, y in points_xy]


def _failed_metrics(error: str) -> dict[str, Any]:
    return {
        "has_solution": False,
        "plan_error": str(error),
        "count_ok": 0,
        "count_collision_error": 0,
        "count_max_speed_error": 0,
        "count_max_accel_error": 0,
        "count_max_curvature_error": 0,
        "trajectory_cost": None,
        "accel_at_1": None,
        "min_accel": None,
        "max_abs_accel": None,
        "max_abs_jerk": None,
        "final_speed_mps": None,
        "min_speed_mps": None,
        "lateral_shift_m": None,
        "final_d_m": None,
        "max_abs_d_m": None,
        "min_dist_to_target_prediction_m": None,
        "ttc_robot_radius_s": None,
        "ttc_2x_robot_radius_s": None,
        "longitudinal_command": "plan_error",
        "lateral_command": "plan_error",
    }


def _infer_commands(metrics: dict[str, Any]) -> tuple[str, str]:
    if not metrics.get("has_solution"):
        return "no_plan", "no_plan"
    accel = float(metrics.get("accel_at_1", 0.0) or 0.0)
    if accel > 0.3:
        long_cmd = "accelerate"
    elif accel < -0.3:
        long_cmd = "brake"
    else:
        long_cmd = "keep_speed"

    lateral_shift = float(metrics.get("lateral_shift_m", 0.0) or 0.0)
    if lateral_shift > 0.8:
        lat_cmd = "lane_change_left"
    elif lateral_shift < -0.8:
        lat_cmd = "lane_change_right"
    else:
        lat_cmd = "keep_lane"
    return long_cmd, lat_cmd


def _trajectory_metrics(
    result: Any,
    *,
    target_points_xy: list[tuple[float, float]],
    robot_radius: float,
    dt: float,
) -> dict[str, Any]:
    counts = dict(result.candidate_counts)
    metrics: dict[str, Any] = {
        "has_solution": bool(result.has_solution),
        "plan_error": "",
        "count_ok": int(counts.get("ok", 0)),
        "count_collision_error": int(counts.get("collision_error", 0)),
        "count_max_speed_error": int(counts.get("max_speed_error", 0)),
        "count_max_accel_error": int(counts.get("max_accel_error", 0)),
        "count_max_curvature_error": int(counts.get("max_curvature_error", 0)),
        "trajectory_cost": None,
        "accel_at_1": None,
        "min_accel": None,
        "max_abs_accel": None,
        "max_abs_jerk": None,
        "final_speed_mps": None,
        "min_speed_mps": None,
        "lateral_shift_m": None,
        "final_d_m": None,
        "max_abs_d_m": None,
        "min_dist_to_target_prediction_m": None,
        "ttc_robot_radius_s": None,
        "ttc_2x_robot_radius_s": None,
    }
    traj = result.best_trajectory
    if traj is None:
        long_cmd, lat_cmd = _infer_commands(metrics)
        metrics["longitudinal_command"] = long_cmd
        metrics["lateral_command"] = lat_cmd
        return metrics

    metrics["trajectory_cost"] = float(traj.cost)
    acc = [float(v) for v in traj.acceleration]
    speed = [float(v) for v in traj.speed]
    d_values = [float(v) for v in traj.d]
    metrics["accel_at_1"] = acc[min(1, len(acc) - 1)] if acc else 0.0
    metrics["min_accel"] = min(acc) if acc else 0.0
    metrics["max_abs_accel"] = max((abs(v) for v in acc), default=0.0)
    if len(acc) >= 2 and dt > 0.0:
        metrics["max_abs_jerk"] = max(abs(acc[i + 1] - acc[i]) / dt for i in range(len(acc) - 1))
    else:
        metrics["max_abs_jerk"] = 0.0
    metrics["final_speed_mps"] = speed[-1] if speed else 0.0
    metrics["min_speed_mps"] = min(speed) if speed else 0.0
    metrics["lateral_shift_m"] = (d_values[-1] - d_values[0]) if len(d_values) >= 2 else 0.0
    metrics["final_d_m"] = d_values[-1] if d_values else 0.0
    metrics["max_abs_d_m"] = max((abs(v) for v in d_values), default=0.0)

    if target_points_xy:
        min_dist = float("inf")
        ttc_r = None
        ttc_2r = None
        for idx, (x, y) in enumerate(zip(traj.x, traj.y)):
            dist = min(math.hypot(float(x) - px, float(y) - py) for px, py in target_points_xy)
            min_dist = min(min_dist, dist)
            t_value = float(traj.t[idx]) if idx < len(traj.t) else idx * dt
            if ttc_r is None and dist <= robot_radius:
                ttc_r = t_value
            if ttc_2r is None and dist <= 2.0 * robot_radius:
                ttc_2r = t_value
        metrics["min_dist_to_target_prediction_m"] = min_dist
        metrics["ttc_robot_radius_s"] = ttc_r
        metrics["ttc_2x_robot_radius_s"] = ttc_2r

    long_cmd, lat_cmd = _infer_commands(metrics)
    metrics["longitudinal_command"] = long_cmd
    metrics["lateral_command"] = lat_cmd
    return metrics


def _degradation_score(clean: dict[str, Any], attacked: dict[str, Any]) -> dict[str, Any]:
    no_solution = int(bool(clean.get("has_solution")) and not bool(attacked.get("has_solution")))
    ok_drop = max(0.0, float(clean.get("count_ok", 0)) - float(attacked.get("count_ok", 0)))
    collision_inc = max(
        0.0,
        float(attacked.get("count_collision_error", 0)) - float(clean.get("count_collision_error", 0)),
    )
    clean_cost = clean.get("trajectory_cost")
    attacked_cost = attacked.get("trajectory_cost")
    cost_inc = (
        max(0.0, float(attacked_cost) - float(clean_cost))
        if clean_cost is not None and attacked_cost is not None
        else 0.0
    )
    clean_min_dist = clean.get("min_dist_to_target_prediction_m")
    attacked_min_dist = attacked.get("min_dist_to_target_prediction_m")
    min_dist_drop = (
        max(0.0, float(clean_min_dist) - float(attacked_min_dist))
        if clean_min_dist is not None and attacked_min_dist is not None
        else 0.0
    )
    clean_min_accel = clean.get("min_accel")
    attacked_min_accel = attacked.get("min_accel")
    decel_inc = (
        max(0.0, -float(attacked_min_accel) - (-float(clean_min_accel)))
        if clean_min_accel is not None and attacked_min_accel is not None
        else 0.0
    )
    lateral_change = abs(
        float(attacked.get("lateral_shift_m") or 0.0) - float(clean.get("lateral_shift_m") or 0.0)
    )
    command_changed = int(
        clean.get("longitudinal_command") != attacked.get("longitudinal_command")
        or clean.get("lateral_command") != attacked.get("lateral_command")
    )
    rank_score = (
        no_solution * 10000.0
        + ok_drop * 20.0
        + collision_inc * 5.0
        + cost_inc * 10.0
        + min_dist_drop * 50.0
        + decel_inc * 100.0
        + lateral_change * 50.0
        + command_changed * 100.0
    )
    return {
        "rank_score": rank_score,
        "no_solution_delta": no_solution,
        "ok_drop": ok_drop,
        "collision_error_increase": collision_inc,
        "cost_increase": cost_inc,
        "min_dist_drop_m": min_dist_drop,
        "decel_increase_mps2": decel_inc,
        "lateral_shift_change_m": lateral_change,
        "command_changed": command_changed,
    }


def _run_combo_task(task: dict[str, Any]) -> dict[str, Any]:
    assert _WORKER_PLANNER_CONFIG is not None
    planner_config = _WORKER_PLANNER_CONFIG
    sample_name = task["sample_name"]
    forward_pct = float(task["forward_pct"])
    lateral_pct = 100.0 - forward_pct
    interval_m = float(task["interval_m"])
    prediction_steps = int(task["prediction_steps"])
    frame_packets = task["frame_packets"]

    frame_rows: list[dict[str, Any]] = []
    for packet in frame_packets:
        ego_state = CartesianEgoState(**packet["ego_state"])
        current_xy = tuple(packet["target_current_xy"])
        lateral_sign = float(packet["lateral_sign"])
        forward_weight = forward_pct / 100.0
        lateral_weight = lateral_pct / 100.0
        direction_xy = _ego_direction_to_global(
            ego=ego_state,
            forward_weight=forward_weight,
            lateral_weight=lateral_weight,
            lateral_sign=lateral_sign,
        )
        fake_points = _fake_future_points(
            current_xy=current_xy,
            direction_xy=direction_xy,
            interval_m=interval_m,
            prediction_steps=prediction_steps,
        )
        obstacles = [
            ObstaclePoint(x=float(x), y=float(y))
            for x, y in packet["tracking_obstacle_xy"]
        ] + _prediction_points_to_obstacles(fake_points)
        try:
            result = _plan_once_with_timeout(
                packet["reference_waypoint_x"],
                packet["reference_waypoint_y"],
                obstacles=obstacles,
                ego_cartesian_state=ego_state,
                config=planner_config,
                timeout_s=float(_WORKER_PLAN_TIMEOUT_S),
            )
            attacked = _trajectory_metrics(
                result,
                target_points_xy=fake_points,
                robot_radius=float(packet["robot_radius"]),
                dt=float(packet["dt"]),
            )
        except Exception as exc:
            attacked = _failed_metrics(f"{type(exc).__name__}: {exc}")
        degradation = _degradation_score(packet["clean_metrics"], attacked)
        row = {
            "sample": sample_name,
            "frame_id": int(packet["frame_id"]),
            "sample_token": packet["sample_token"],
            "forward_pct": forward_pct,
            "lateral_toward_ego_pct": lateral_pct,
            "interval_m": interval_m,
            "prediction_steps": prediction_steps,
            "target_ego_front_m": float(packet["target_ego_front_m"]),
            "target_ego_left_m": float(packet["target_ego_left_m"]),
            "lateral_sign": lateral_sign,
            **{f"attack_{k}": v for k, v in attacked.items()},
            **degradation,
        }
        frame_rows.append(row)

    aggregate = {
        "sample": sample_name,
        "forward_pct": forward_pct,
        "lateral_toward_ego_pct": lateral_pct,
        "interval_m": interval_m,
        "prediction_steps": prediction_steps,
        "rank_score_sum": sum(float(row["rank_score"]) for row in frame_rows),
        "rank_score_mean": sum(float(row["rank_score"]) for row in frame_rows) / max(1, len(frame_rows)),
        "rank_score_max": max((float(row["rank_score"]) for row in frame_rows), default=0.0),
        "no_solution_frames": sum(int(row["no_solution_delta"]) for row in frame_rows),
        "command_changed_frames": sum(int(row["command_changed"]) for row in frame_rows),
        "ok_drop_sum": sum(float(row["ok_drop"]) for row in frame_rows),
        "collision_error_increase_sum": sum(float(row["collision_error_increase"]) for row in frame_rows),
        "cost_increase_sum": sum(float(row["cost_increase"]) for row in frame_rows),
        "max_decel_increase_mps2": max((float(row["decel_increase_mps2"]) for row in frame_rows), default=0.0),
        "max_lateral_shift_change_m": max((float(row["lateral_shift_change_m"]) for row in frame_rows), default=0.0),
        "min_attack_dist_to_target_prediction_m": min(
            (
                float(row["attack_min_dist_to_target_prediction_m"])
                for row in frame_rows
                if row.get("attack_min_dist_to_target_prediction_m") is not None
            ),
            default=None,
        ),
        "frame_rows": frame_rows,
    }
    return aggregate


def _append_dict_rows(sheet: Any, rows: list[dict[str, Any]], columns: list[str]) -> None:
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(col) for col in columns])


def _autosize(sheet: Any) -> None:
    for column_cells in sheet.columns:
        max_len = 0
        letter = column_cells[0].column_letter
        for cell in column_cells[:200]:
            value = cell.value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        sheet.column_dimensions[letter].width = min(max(10, max_len + 2), 40)


def _write_xlsx(
    *,
    output_path: Path,
    clean_rows: list[dict[str, Any]],
    combo_rows: list[dict[str, Any]],
    frame_rows: list[dict[str, Any]],
    top_k: int,
    case_by_sample: dict[str, str] | None = None,
    case_sheets_only: bool = False,
    error_rows: list[dict[str, Any]] | None = None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    top_rows: list[dict[str, Any]] = []
    top_frame_rows: list[dict[str, Any]] = []
    for sample_name in sorted({row["sample"] for row in combo_rows}):
        sample_rows = [row for row in combo_rows if row["sample"] == sample_name]
        sample_rows.sort(key=lambda row: (-float(row["rank_score_sum"]), float(row["interval_m"])))
        for rank, row in enumerate(sample_rows[:top_k], start=1):
            out = dict(row)
            out["case"] = (case_by_sample or {}).get(str(sample_name), "")
            out["sample_rank"] = rank
            out.pop("frame_rows", None)
            top_rows.append(out)
            for frame_row in row["frame_rows"]:
                frame_out = dict(frame_row)
                frame_out["case"] = (case_by_sample or {}).get(str(sample_name), "")
                frame_out["sample_rank"] = rank
                top_frame_rows.append(frame_out)

    all_combo_rows = []
    for row in combo_rows:
        out = dict(row)
        out["case"] = (case_by_sample or {}).get(str(row.get("sample", "")), "")
        out.pop("frame_rows", None)
        all_combo_rows.append(out)

    if case_sheets_only:
        for case_name in sorted(set((case_by_sample or {}).values())):
            rows = [row for row in top_rows if row.get("case") == case_name]
            rows.sort(key=lambda row: str(row.get("sample", "")))
            ws = wb.create_sheet(case_name[:31])
            if rows:
                columns = list(rows[0].keys())
                _append_dict_rows(ws, rows, columns)
            _autosize(ws)
        unknown_rows = [row for row in top_rows if not row.get("case")]
        if unknown_rows:
            ws = wb.create_sheet("unknown_case")
            _append_dict_rows(ws, unknown_rows, list(unknown_rows[0].keys()))
            _autosize(ws)
        if error_rows:
            ws = wb.create_sheet("errors")
            _append_dict_rows(ws, error_rows, list(error_rows[0].keys()))
            _autosize(ws)
        wb.save(output_path)
        return

    sheets = [
        ("top10_by_sample", top_rows),
        ("top10_frame_details", top_frame_rows),
        ("clean_baseline", clean_rows),
        ("all_combos", all_combo_rows),
        ("all_frame_results", frame_rows),
    ]
    for name, rows in sheets:
        ws = wb.create_sheet(name[:31])
        if rows:
            columns = list(rows[0].keys())
            _append_dict_rows(ws, rows, columns)
        _autosize(ws)
    if error_rows:
        ws = wb.create_sheet("errors")
        _append_dict_rows(ws, error_rows, list(error_rows[0].keys()))
        _autosize(ws)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Sweep forged target-car predictions through zz7 FOT planning")
    parser.add_argument("--config", type=Path, default=PROJECT_ROOT / "config.yaml")
    parser.add_argument("--cache-root", type=Path, default=CAMOUFLAGE_ROOT / "data-pre")
    parser.add_argument("--yaml-root", type=Path, default=CAMOUFLAGE_ROOT / "data-yaml")
    parser.add_argument("--match-config", type=Path, default=CAMOUFLAGE_ROOT / "config.yaml")
    parser.add_argument("--scenario-root", type=Path, default=None)
    parser.add_argument("--samples", default=",".join(DEFAULT_SAMPLES))
    parser.add_argument("--intervals", default="0.5,1.0,1.5,2.0,2.5,3.0,4.0,5.0,6.0,8.0")
    parser.add_argument("--direction-step-pct", type=int, default=1)
    parser.add_argument("--prediction-steps", type=int, default=8)
    parser.add_argument("--top-k", type=int, default=10)
    parser.add_argument("--workers", type=int, default=max(1, min(8, os.cpu_count() or 1)))
    parser.add_argument("--binding-workers", type=int, default=max(1, min(8, os.cpu_count() or 1)))
    parser.add_argument("--plan-timeout-s", type=float, default=10.0)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--case-sheets-only", action="store_true")
    args = parser.parse_args()

    if args.scenario_root is not None:
        cases = _discover_scenario_cases(args.scenario_root)
        sample_names = sorted({sample for samples in cases.values() for sample in samples})
        case_by_sample = {
            sample: case_name for case_name, samples in cases.items() for sample in samples
        }
    else:
        sample_names = [part.strip() for part in args.samples.split(",") if part.strip()]
        case_by_sample = {}
    intervals = [float(part.strip()) for part in args.intervals.split(",") if part.strip()]
    direction_values = list(range(0, 101, max(1, int(args.direction_step_pct))))
    if direction_values[-1] != 100:
        direction_values.append(100)

    config = _load_config(args.config)
    paths_cfg = config.get("paths", {})
    planner_cfg_values = dict(config.get("planner", {}))
    adapter_cfg_values = dict(config.get("adapter", {}))
    adapter_cfg_values.update(
        {
            "include_tracking_boxes": True,
            "include_prediction_points": False,
            "prediction_mode": "none",
        }
    )
    planner_config = FrenetPlannerConfig(**planner_cfg_values)
    adapter_config = NuScenesFOTAdapterConfig(**adapter_cfg_values)

    binding_errors = _ensure_binding_cache_parallel(
        sample_names,
        cache_root=args.cache_root,
        yaml_root=args.yaml_root,
        match_config_path=args.match_config,
        workers=int(args.binding_workers),
    )
    binding_error_by_sample = {row["sample_name"]: row for row in binding_errors}
    sample_names = [
        sample_name
        for sample_name in sample_names
        if sample_name not in binding_error_by_sample
        and list((args.cache_root / sample_name).glob("cache-*/target-car-binding.yaml"))
    ]
    if binding_errors:
        print(f"[binding] skipped samples with binding errors={len(binding_errors)}", flush=True)
    error_rows = [
        {
            "case": case_by_sample.get(str(row.get("sample_name", "")), ""),
            "sample": row.get("sample_name", ""),
            "stage": "target_binding",
            "error": row.get("error", ""),
            "path": row.get("path", ""),
        }
        for row in binding_errors
    ]
    bindings = _load_bindings(
        sample_names,
        args.cache_root,
        yaml_root=args.yaml_root,
        match_config_path=args.match_config,
    )
    sample_tokens = [
        str(frame["sample_token"])
        for binding in bindings.values()
        for frame in binding.get("frames", [])
    ]
    gt_tracking_path = args.output.with_suffix(".gt_tracking.json")
    _build_gt_tracking_json(
        nuscenes_root=Path(paths_cfg["nuscenes_root"]),
        sample_tokens=sample_tokens,
        output_path=gt_tracking_path,
    )

    adapter = NuScenesFOTAdapter(
        NuScenesFOTPaths(
            nuscenes_root=str(paths_cfg["nuscenes_root"]),
            tracking_results_path=str(gt_tracking_path),
            prediction_results_path=None,
        )
    )

    clean_rows: list[dict[str, Any]] = []
    frame_packets_by_sample: dict[str, list[dict[str, Any]]] = {}
    for sample_name, binding in bindings.items():
        frames = sorted(binding.get("frames", []), key=lambda row: int(row["frame_id"]))
        centers = [
            (float(frame["gt"]["center_xyz"][0]), float(frame["gt"]["center_xyz"][1]))
            for frame in frames
        ]
        packets: list[dict[str, Any]] = []
        for frame_index, frame in enumerate(frames):
            sample_token = str(frame["sample_token"])
            prepared = adapter.prepare_input(
                sample_token,
                adapter_config=adapter_config,
                planner_dt=float(planner_config.dt),
            )
            target_current_xy = centers[frame_index]
            clean_points = _extend_clean_future_points(
                centers,
                frame_index,
                prediction_steps=int(args.prediction_steps),
            )
            clean_obstacles = list(prepared.obstacle_points) + _prediction_points_to_obstacles(clean_points)
            try:
                clean_result = _plan_once_with_timeout(
                    prepared.reference_waypoint_x,
                    prepared.reference_waypoint_y,
                    obstacles=clean_obstacles,
                    ego_cartesian_state=prepared.ego_state,
                    config=planner_config,
                    timeout_s=float(args.plan_timeout_s),
                )
                clean_metrics = _trajectory_metrics(
                    clean_result,
                    target_points_xy=clean_points,
                    robot_radius=float(planner_config.robot_radius),
                    dt=float(planner_config.dt),
                )
            except Exception as exc:
                clean_metrics = _failed_metrics(f"{type(exc).__name__}: {exc}")
            target_front, target_left = _ego_local_xy(target_current_xy, prepared.ego_state)
            lateral_sign = -1.0 if target_left > 0.0 else 1.0
            clean_row = {
                "case": case_by_sample.get(sample_name, ""),
                "sample": sample_name,
                "frame_id": int(frame["frame_id"]),
                "sample_token": sample_token,
                "binding_path": binding.get("_binding_path", ""),
                "target_instance_token": str(frame.get("target_instance_token", "")),
                "target_ann_token": str(frame.get("ann_token", "")),
                "target_global_x": target_current_xy[0],
                "target_global_y": target_current_xy[1],
                "target_ego_front_m": target_front,
                "target_ego_left_m": target_left,
                "lateral_sign_toward_ego": lateral_sign,
                **{f"clean_{key}": value for key, value in clean_metrics.items()},
            }
            clean_rows.append(clean_row)
            packets.append(
                {
                    "sample_name": sample_name,
                    "frame_id": int(frame["frame_id"]),
                    "sample_token": sample_token,
                    "ego_state": asdict(prepared.ego_state),
                    "reference_waypoint_x": list(prepared.reference_waypoint_x),
                    "reference_waypoint_y": list(prepared.reference_waypoint_y),
                    "tracking_obstacle_xy": [
                        (float(point.x), float(point.y)) for point in prepared.tracking_obstacle_points
                    ],
                    "target_current_xy": target_current_xy,
                    "target_ego_front_m": target_front,
                    "target_ego_left_m": target_left,
                    "lateral_sign": lateral_sign,
                    "clean_metrics": clean_metrics,
                    "robot_radius": float(planner_config.robot_radius),
                    "dt": float(planner_config.dt),
                }
            )
        frame_packets_by_sample[sample_name] = packets

    tasks: list[dict[str, Any]] = []
    for sample_name, packets in frame_packets_by_sample.items():
        for forward_pct in direction_values:
            for interval_m in intervals:
                tasks.append(
                    {
                        "sample_name": sample_name,
                        "forward_pct": float(forward_pct),
                        "interval_m": float(interval_m),
                        "prediction_steps": int(args.prediction_steps),
                        "frame_packets": packets,
                    }
                )

    start = time.time()
    combo_rows: list[dict[str, Any]] = []
    frame_rows: list[dict[str, Any]] = []
    print(
        f"[sweep] samples={len(sample_names)} tasks={len(tasks)} "
        f"frame_plans={len(tasks) * 3} workers={args.workers} output={args.output}",
        flush=True,
    )
    if args.workers <= 1:
        _worker_init(planner_cfg_values, float(args.plan_timeout_s))
        iterator = map(_run_combo_task, tasks)
        for index, result in enumerate(iterator, start=1):
            combo_rows.append(result)
            frame_rows.extend(result["frame_rows"])
            if index % 100 == 0 or index == len(tasks):
                elapsed = time.time() - start
                print(f"[sweep] {index}/{len(tasks)} elapsed={elapsed:.1f}s", flush=True)
    else:
        with mp.Pool(
            processes=int(args.workers),
            initializer=_worker_init,
            initargs=(planner_cfg_values, float(args.plan_timeout_s)),
        ) as pool:
            for index, result in enumerate(pool.imap_unordered(_run_combo_task, tasks, chunksize=4), start=1):
                combo_rows.append(result)
                frame_rows.extend(result["frame_rows"])
                if index % 100 == 0 or index == len(tasks):
                    elapsed = time.time() - start
                    print(f"[sweep] {index}/{len(tasks)} elapsed={elapsed:.1f}s", flush=True)

    combo_rows.sort(key=lambda row: (row["sample"], -float(row["rank_score_sum"])))
    frame_rows.sort(key=lambda row: (row["sample"], row["frame_id"], -float(row["rank_score"])))
    _write_xlsx(
        output_path=args.output,
        clean_rows=clean_rows,
        combo_rows=combo_rows,
        frame_rows=frame_rows,
        top_k=int(args.top_k),
        case_by_sample=case_by_sample,
        case_sheets_only=bool(args.case_sheets_only),
        error_rows=error_rows,
    )
    print(f"[sweep] done output={args.output} elapsed={time.time() - start:.1f}s", flush=True)
    print(f"[sweep] gt_tracking={gt_tracking_path}", flush=True)


if __name__ == "__main__":
    main()
